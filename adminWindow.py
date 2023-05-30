# pyQT5 libraries
from PyQt5.QtGui import QImage, QPixmap, QColor
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QApplication, QDialog, QMainWindow, QMessageBox, QWidget, QFileDialog
from PyQt5.QtCore import pyqtSlot, QTimer, Qt
from PyQt5.uic import loadUi

# library for handling excel automation
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

import sqlite3
import dateutil.relativedelta as rdelta
import datetime
import cv2
import sys
import face_recognition
import numpy as np
import os
import bcrypt
import tkinter
from tkinter import filedialog

# set up system path if run as python file or as exe file
if getattr(sys, 'frozen', False):
    # If the application is run as a bundle, the PyInstaller bootloader
    # extends the sys module by a flag frozen=True and sets the app
    # path into variable _MEIPASS'.
    application_path = sys._MEIPASS
else:
    application_path = os.path.dirname(os.path.abspath(__file__))


class AdminWindow(QMainWindow):
    def __init__(self):
        # call parent constructor
        super(AdminWindow, self).__init__()

        # load user interface for admin window
        loadUi("adminwindowSmaller.ui", self)

        # start maximized
        self.setWindowState(Qt.WindowMaximized)

        # message box for displaying errors
        self.msgbox = QMessageBox()
        self.msgbox.setWindowTitle("warning")
        self.msgbox.setIcon(QMessageBox.Warning)

        # message box for asking confirmation
        self.confirm_msgbox = QMessageBox()

        # video variable for taking picture
        self.capture = None

        # set up path to database for exe
        database_path = os.path.join(application_path, 'attendance.db')

        # start a connection to sql database
        self.conn = None
        self.conn = sqlite3.connect(database_path)

        # create sql cursor
        self.c = self.conn.cursor()

        # load face_encodings
        self.c.execute("SELECT face_encoding FROM employees")
        rows = self.c.fetchall()
        self.face_encodings = []

        for row in rows:
            self.face_encodings.append(np.array(eval(row[0])))


        # navigation buttons
        self.homeButton.clicked.connect(lambda: self.switchPage(0))
        self.empListButton.clicked.connect(lambda: self.switchPage(1))
        self.managerButton.clicked.connect(lambda: self.switchPage(2))
        self.adminButton.clicked.connect(lambda: self.switchPage(3))
        self.exportButton.clicked.connect(self.exportAttendance)
        self.logoutButton.clicked.connect(self.close)

        # other buttons
        self.captureButton.clicked.connect(self.captureImage)
        self.uploadButton.clicked.connect(self.uploadImageFunction)
        self.addEmployeeButton.clicked.connect(self.addEmployeeFunction)

        # add buttons
        self.addManagerButton.clicked.connect(self.addMngrButton)
        self.registerAdminButton.clicked.connect(self.addAdminButton)

        # remove buttons
        self.removeEmpButton.clicked.connect(lambda: self.removeEmployeeButton())
        self.removeMngrButton.clicked.connect(self.removeManagerButton)
        self.removeAdminButton.clicked.connect(self.removeAdmnButton)

        # add and remove buttons for each page to open add and remove pages
        self.empListAdd.clicked.connect(self.addEmployeePage)
        self.mngrListAdd.clicked.connect(self.addManagerPage)
        self.adminListAdd.clicked.connect(self.addAdminPage)
        self.empListRemove.clicked.connect(self.removeEmployeePage)
        self.mngrListRemove.clicked.connect(self.removeManagerPage)
        self.adminListRemove.clicked.connect(self.removeAdminPage)

        # call homepage to set up table
        self.switchPage(0)



    def switchPage(self, page):
        # stackedWidget is for table pages, stackedWidget_2 is for page on the right to add and remove
        self.stackedWidget.setCurrentIndex(page)
        self.stackedWidget_2.setCurrentIndex(0)

        # release camera if left opened
        if self.capture is not None:
            if self.capture.isOpened():
                self.capture.release()
                # stop updating frames
                self.timer.stop()

        if page == 0:
            # set headers to visible because switching pages resets setVisible to false
            self.homeTable.verticalHeader().setVisible(True)
            self.homeTable.horizontalHeader().setVisible(True)

            now = datetime.datetime.now()

            date_string = now.strftime("%Y-%m-%d")

            # execute query to get all employee information
            self.c.execute("""SELECT employees.name, employees.employee_ID, time_in, time_out, total_time
                                        FROM employees 
                                        INNER JOIN attendance ON employees.employee_ID = attendance.employee_ID
                                        WHERE date = ?
                                    """, (date_string,))
        elif page == 1:
            # set headers to visible because switching pages resets setVisible to false
            self.empTable.verticalHeader().setVisible(True)
            self.empTable.horizontalHeader().setVisible(True)

            # execute query to get all employee information
            self.c.execute("""SELECT employees.employee_ID, employees.name, departments.name
                                                FROM employees 
                                                INNER JOIN departments ON employees.department_ID = departments.department_ID
                                            """)
        elif page == 2:
            # set headers to visible because switching pages resets setVisible to false
            self.mngrTable.verticalHeader().setVisible(True)
            self.mngrTable.horizontalHeader().setVisible(True)

            # execute query to get all employee information
            self.c.execute("""SELECT employees.employee_ID, employees.name, departments.name
                                       FROM employees 
                                       INNER JOIN hr_managers ON employees.employee_ID = hr_managers.employee_ID
                                       INNER JOIN departments ON employees.department_ID = departments.department_ID
                                   """)
        elif page == 3:
            # set headers to visible because switching pages resets setVisible to false
            self.adminTable.verticalHeader().setVisible(True)
            self.adminTable.horizontalHeader().setVisible(True)

            # execute query to get all employee information
            self.c.execute("""SELECT employees.employee_ID, employees.name, departments.name
                                                                   FROM employees 
                                                                   INNER JOIN hr_managers ON employees.employee_ID = hr_managers.employee_ID
                                                                   INNER JOIN departments ON employees.department_ID = departments.department_ID
                                                                   INNER JOIN admins ON hr_managers.manager_ID = admins.manager_ID
                                                               """)

        # fetch results
        rows = self.c.fetchall()

        # display table items depending on table
        if page == 0:
            # display items to table
            self.homeTable.setRowCount(len(rows))
            row_count = 0
            for row in rows:
                self.homeTable.setItem(row_count, 0, QtWidgets.QTableWidgetItem(row[0]))
                self.homeTable.setItem(row_count, 1, QtWidgets.QTableWidgetItem(row[1]))
                self.homeTable.setItem(row_count, 2, QtWidgets.QTableWidgetItem(row[2]))
                self.homeTable.setItem(row_count, 3, QtWidgets.QTableWidgetItem(row[3]))
                self.homeTable.setItem(row_count, 4, QtWidgets.QTableWidgetItem(row[4]))
                row_count += 1
        elif page == 1:
            # display items to table
            self.empTable.setRowCount(len(rows))
            row_count = 0
            for row in rows:
                self.empTable.setItem(row_count, 0, QtWidgets.QTableWidgetItem(row[0]))
                self.empTable.setItem(row_count, 1, QtWidgets.QTableWidgetItem(row[1]))
                self.empTable.setItem(row_count, 2, QtWidgets.QTableWidgetItem(row[2]))
                row_count += 1
        elif page == 2:
            # display items to table
            self.mngrTable.setRowCount(len(rows))
            row_count = 0
            for row in rows:
                self.mngrTable.setItem(row_count, 0, QtWidgets.QTableWidgetItem(row[0]))
                self.mngrTable.setItem(row_count, 1, QtWidgets.QTableWidgetItem(row[1]))
                self.mngrTable.setItem(row_count, 2, QtWidgets.QTableWidgetItem(row[2]))
                row_count += 1
        elif page == 3:
            # display items to table
            self.adminTable.setRowCount(len(rows))
            row_count = 0
            for row in rows:
                self.adminTable.setItem(row_count, 0, QtWidgets.QTableWidgetItem(row[0]))
                self.adminTable.setItem(row_count, 1, QtWidgets.QTableWidgetItem(row[1]))
                self.adminTable.setItem(row_count, 2, QtWidgets.QTableWidgetItem(row[2]))
                row_count += 1

    def addEmployeePage(self):
        # switch to page
        self.stackedWidget_2.setCurrentIndex(1)

        # release videocapture variable if used before
        if self.capture is not None:
            if self.capture.isOpened():
                self.capture.release()

        # empty form
        self.nameEdit.setText("")
        self.idEdit.setText("")
        self.deptEdit.setText("")

        # reset image
        default_picture = cv2.imread(os.path.join(application_path, "icons/profile.png"))

        # converts picture to proper rgb format for displaying in window
        qformat = QImage.Format_Indexed8
        if len(default_picture.shape) == 3:
            if default_picture.shape[2] == 4:
                qformat = QImage.Format_RGBA8888
            else:
                qformat = QImage.Format_RGB888
        default_picture = QImage(default_picture, default_picture.shape[1], default_picture.shape[0],
                                 default_picture.strides[0], qformat)

        # convert rgb to bgr for proper output
        default_picture = default_picture.rgbSwapped()

        self.picLabel.setPixmap(QPixmap.fromImage(default_picture))

    def uploadImageFunction(self):
        # release camera if left opened
        if self.capture is not None:
            if self.capture.isOpened():
                self.capture.release()
                self.timer.stop()

        self.stackedWidget.setCurrentIndex(4)

        self.current_frame = None

        # call methods
        self.startVideo()

    def addEmployeeFunction(self):
        # disable add button
        self.addEmployeeButton.setEnabled(False)

        name = self.nameEdit.text()
        employeeID = self.idEdit.text()
        department = self.deptEdit.text()

        cvpictureExists = None

        self.c.execute("SELECT * FROM employees WHERE employee_ID = ?", (employeeID,))
        rows = self.c.fetchall()

        alreadyInDatabase = None

        if not rows:
            alreadyInDatabase = False
        else:
            alreadyInDatabase = True

        alreadyHasFaceEncoding = False

        try:
            if self.cvpicture.any():
                cvpictureExists = True

                # resize the frame to a smaller size to make computations faster
                img_small = cv2.resize(self.cvpicture, (0, 0), None, 0.25, 0.25)

                # convert to rgb
                img_small = cv2.cvtColor(img_small, cv2.COLOR_BGR2RGB)

                # find locations of faces in the image
                faces_cur_frame = face_recognition.face_locations(img_small)

                encodings_cur_frame = face_recognition.face_encodings(img_small, faces_cur_frame)

                if self.face_encodings:
                    for encoding in encodings_cur_frame:
                        match = face_recognition.compare_faces(self.face_encodings, encoding, tolerance=0.50)
                        face_dis = face_recognition.face_distance(self.face_encodings, encoding)
                        best_match_index = np.argmin(face_dis)

                        # pass in best_match_index, -1 if found face was unknown
                        if match[best_match_index]:
                            alreadyHasFaceEncoding = True

        except Exception:
            cvpictureExists = False



        if not name:
            self.msgbox.setText("must provide name")
            self.msgbox.exec_()
        elif not employeeID:
            self.msgbox.setText("must provide employeeID")
            self.msgbox.exec_()
        elif not department:
            self.msgbox.setText("must provide department")
            self.msgbox.exec_()
        elif not cvpictureExists:
            self.msgbox.setText("must provide a picture")
            self.msgbox.exec_()
        elif alreadyInDatabase:
            self.msgbox.setText(f"Error: Employee ID already Taken")
            self.msgbox.exec_()
            # empty form
            self.nameEdit.setText("")
            self.idEdit.setText("")
            self.deptEdit.setText("")
        elif alreadyHasFaceEncoding:
            self.msgbox.setText("Error: Employee with similar face encodings already in the database")
            self.msgbox.exec_()

            # reset image
            default_picture = cv2.imread(os.path.join(application_path, "icons/profile.png"))

            # converts picture to proper rgb format for displaying in window
            qformat = QImage.Format_Indexed8
            if len(default_picture.shape) == 3:
                if default_picture.shape[2] == 4:
                    qformat = QImage.Format_RGBA8888
                else:
                    qformat = QImage.Format_RGB888
            default_picture = QImage(default_picture, default_picture.shape[1], default_picture.shape[0],
                                     default_picture.strides[0], qformat)

            # convert rgb to bgr for proper output
            default_picture = default_picture.rgbSwapped()

            self.picLabel.setPixmap(QPixmap.fromImage(default_picture))

            # empty form
            self.nameEdit.setText("")
            self.idEdit.setText("")
            self.deptEdit.setText("")
        else:
            # insert department if new
            self.c.execute("INSERT OR IGNORE INTO departments (name) VALUES(?)", (department,))
            self.conn.commit()

            self.c.execute("SELECT department_ID from departments WHERE name = ?", (department,))
            rows = self.c.fetchall()

            # save image to local folder
            filename = f"ImagesAttendance/{employeeID}.jpg"
            cv2.imwrite(os.path.join(application_path, filename), self.cvpicture)

            imagePath = f"ImagesAttendance/{employeeID}.jpg"

            # insert employee to database
            self.c.execute("""INSERT INTO employees (employee_ID, department_ID, 
                            name, face_image_path, face_encoding) VALUES(?,?,?,?,?)""",
                           (employeeID, rows[0][0], name, imagePath, (str(self.face_encoding.tolist()))))
            self.conn.commit()

            # reload table
            self.switchPage(1)

            # reset image
            default_picture = cv2.imread(os.path.join(application_path, "icons/profile.png"))

            # converts picture to proper rgb format for displaying in window
            qformat = QImage.Format_Indexed8
            if len(default_picture.shape) == 3:
                if default_picture.shape[2] == 4:
                    qformat = QImage.Format_RGBA8888
                else:
                    qformat = QImage.Format_RGB888
            default_picture = QImage(default_picture, default_picture.shape[1], default_picture.shape[0],
                                     default_picture.strides[0], qformat)

            # convert rgb to bgr for proper output
            default_picture = default_picture.rgbSwapped()

            self.picLabel.setPixmap(QPixmap.fromImage(default_picture))

            # empty form
            self.nameEdit.setText("")
            self.idEdit.setText("")
            self.deptEdit.setText("")

        # enable add button
        self.addEmployeeButton.setEnabled(True)

    # method to start video capture
    def startVideo(self):
        """
        :return: no return value
        """

        # index for which camera to use (0 for default system camera)
        camera_id = 0

        # Open camera for video capture
        self.capture = cv2.VideoCapture(camera_id)

        # call update frame method
        self.update_frame()

        # set up a timer to update frame every 10 milliseconds
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_frame)
        self.timer.start(1)

    # method for updating video frame
    @pyqtSlot()
    def update_frame(self):
        # get frame from video capture, ret is the return value indicating whether a frame has been returned or not
        ret, self.current_frame = self.capture.read()

        # if the frame exists, set it to instance variable and display the frame in the ui
        if ret:

            # converts frame to proper rgb format
            qformat = QImage.Format_Indexed8
            if len(self.current_frame.shape) == 3:
                if self.current_frame.shape[2] == 4:
                    qformat = QImage.Format_RGBA8888
                else:
                    qformat = QImage.Format_RGB888
            outImage = QImage(self.current_frame, self.current_frame.shape[1], self.current_frame.shape[0],
                              self.current_frame.strides[0], qformat)

            # convert rgb to bgr for proper output
            outImage = outImage.rgbSwapped()
            outImage = outImage.mirrored(horizontal=True, vertical=False)

            # display image to imgLabel in user interface
            self.imgLabel.setPixmap(QPixmap.fromImage(outImage))

    def captureImage(self):
        # stop updating frames
        self.timer.stop()

        # close camera
        self.capture.release()

        # switch back to employee page
        self.stackedWidget.setCurrentIndex(1)

        faceDetected = True

        try:
            self.face_encoding = face_recognition.face_encodings(self.current_frame)[0]
        except IndexError:
            faceDetected = False

        if faceDetected == True:
            # converts frame to proper rgb format
            qformat = QImage.Format_Indexed8
            if len(self.current_frame.shape) == 3:
                if self.current_frame.shape[2] == 4:
                    qformat = QImage.Format_RGBA8888
                else:
                    qformat = QImage.Format_RGB888
            outImage = QImage(self.current_frame, self.current_frame.shape[1], self.current_frame.shape[0],
                              self.current_frame.strides[0], qformat)

            # convert rgb to bgr for proper output
            outImage = outImage.rgbSwapped()
            outImage = outImage.mirrored(horizontal=True, vertical=False)

            # display image to imgLabel in user interface
            self.picLabel.setPixmap(QPixmap.fromImage(outImage))

            # take current frame as picture to save
            self.cvpicture = self.current_frame
        else:
            self.msgbox.setText("face cannot be detected, please take another photo")
            self.msgbox.exec_()

    def removeEmployeePage(self):
        # switch to page
        self.stackedWidget_2.setCurrentIndex(4)

        # empty form
        self.removeEmpLE.setText("")

    def removeEmployeeButton(self):

        employeeID = self.removeEmpLE.text()
        self.msgbox.setText("")

        confirmation = self.confirm_msgbox.question(self, "Confirmation", "Remove employee from the database?", QMessageBox.Yes | QMessageBox.Cancel)
        if confirmation == QMessageBox.Yes:
            if employeeID:
                self.c.execute("SELECT * FROM employees WHERE employee_ID = ?", (employeeID,))
                rows = self.c.fetchall()

                if len(rows) != 0:
                    # remove all employee's attendance
                    self.c.execute("DELETE FROM attendance WHERE employee_ID = ?", (employeeID,))
                    self.conn.commit()

                    # remove from admin list if admin
                    self.c.execute("""DELETE FROM admins WHERE admins.manager_ID 
                                        IN (SELECT hr_managers.manager_ID FROM admins 
                                        JOIN hr_managers ON admins.manager_ID = hr_managers.manager_ID
                                        WHERE employee_ID = ?)""", (employeeID,))
                    self.conn.commit()

                    # remove from manager list if manager
                    self.c.execute("DELETE FROM hr_managers WHERE employee_ID = ?", (employeeID,))
                    self.conn.commit()

                    # remove from employee list
                    self.c.execute("DELETE FROM employees WHERE employee_ID = ?", (employeeID,))
                    self.conn.commit()

                    # remove image of employee from directory
                    filename = f"ImagesAttendance/{employeeID}.jpg"

                    os.remove(os.path.join(application_path, filename))

                else:
                    self.msgbox.setText(f"There is no employee with employee id of {employeeID}")
                    self.msgbox.exec_()

                # reload table
                self.switchPage(1)

                # empty form
                self.removeEmpLE.setText("")

            else:
                self.msgbox.setText("Please provide an employee id to delete from database")
                self.msgbox.exec_()

    def addManagerPage(self):
        # switch to page
        self.stackedWidget_2.setCurrentIndex(2)

        self.addManagerLE.setText("")

    def addMngrButton(self):
        employeeID = self.addManagerLE.text()

        if employeeID:
            self.c.execute("SELECT employee_ID FROM employees WHERE employee_ID = ?", (employeeID,))
            rows = self.c.fetchall()

            if not rows:
                self.msgbox.setText(f"There is no employee with ID of {employeeID}")
                self.msgbox.exec_()
                self.addManagerLE.setText("")
                return
            else:
                self.c.execute("SELECT employee_ID FROM hr_managers WHERE employee_ID = ?", (employeeID,))
                newRows = self.c.fetchall()

                if newRows:
                    self.msgbox.setText(f"Employee with id of {employeeID} is already in the managers list")
                    self.msgbox.exec_()
                else:
                    self.c.execute("INSERT INTO hr_managers (employee_ID) VALUES(?)", (employeeID,))
                    self.conn.commit()
            self.addManagerLE.setText("")
        else:
            self.msgbox.setText("Please enter an employee number to add to manager list")
            self.msgbox.exec_()
            self.addManagerLE.setText("")

        # reset table
        self.switchPage(2)

    def removeManagerPage(self):
        # switch to page
        self.stackedWidget_2.setCurrentIndex(5)

        # clear content
        self.removeMngrLE.setText("")

    def removeManagerButton(self):
        employeeID = self.removeMngrLE.text()

        if employeeID:
            self.c.execute("SELECT * FROM hr_managers WHERE employee_ID = ?", (employeeID,))
            rows = self.c.fetchall()

            if not rows:
                self.msgbox.setText(f"There is no Manager with id of {employeeID}")
                self.msgbox.exec_()
                self.removeMngrLE.setText("")
            else:
                # remove from admin list if admin
                self.c.execute("""DELETE FROM admins WHERE admins.manager_ID 
                                    IN (SELECT hr_managers.manager_ID FROM admins 
                                    JOIN hr_managers ON admins.manager_ID = hr_managers.manager_ID
                                    WHERE employee_ID = ?)""", (employeeID,))
                self.conn.commit()

                self.c.execute("DELETE FROM hr_managers WHERE employee_ID = ?", (employeeID,))
                self.conn.commit()
                self.removeMngrLE.setText("")


        else:
            self.msgbox.setText("Please enter manager ID to remove")
            self.msgbox.exec_()


        # reset table
        self.switchPage(2)

    def addAdminPage(self):
        # switch to page
        self.stackedWidget_2.setCurrentIndex(3)


    def addAdminButton(self):

        # get user input
        employeeID = self.employeeIDLE.text()
        username = self.usernameLE.text()
        password = self.passwordLE.text()
        confirmPass = self.confirmLE.text()

        if not employeeID:
            self.msgbox.setText("Please enter employeeID")
            self.msgbox.exec_()
            return
        elif not username:
            self.msgbox.setText("Please enter username")
            self.msgbox.exec_()
            return
        elif not password:
            self.msgbox.setText("Please enter password")
            self.msgbox.exec_()
            return
        elif not confirmPass:
            self.msgbox.setText("Please confirm password")
            self.msgbox.exec_()
            return
        elif password != confirmPass:
            self.msgbox.setText("Passwords do not match")
            self.msgbox.exec_()
            return

        self.c.execute("""SELECT hr_managers.manager_ID, hr_managers.employee_ID FROM hr_managers 
                    WHERE employee_ID = ? """, (employeeID,))
        rows = self.c.fetchall()

        self.c.execute("SELECT username, manager_ID FROM admins")
        admin_rows = self.c.fetchall()
        username_taken = False
        already_admin = False

        i = 0
        for row in admin_rows:
            if row[0] == username:
                username_taken = True

            if not rows:
                pass
            else:
                if rows[i][0] == row[1]:
                    already_admin = True

            i += 1

        # if employeeID is not in managers list then cancel
        if not rows:
            self.msgbox.setText("""The employee ID entered is not eligible for admin registration. Please ask a manager or an admin for assistance.""")
            self.msgbox.exec_()
            self.clearAddAdmin()
            return
        else:
            if username_taken:
                self.msgbox.setText("Error: Username already taken")
                self.msgbox.exec_()
                self.clearAddAdmin()
            elif already_admin:
                self.msgbox.setText("Error: Employee is already an admin")
                self.msgbox.exec_()
                self.clearAddAdmin()
            else:
                # convert to bytes for hashing
                passwordBytes = password.encode('utf-8')
                # generate password salt
                salt = bcrypt.gensalt()
                # Hashing the password
                hash = bcrypt.hashpw(passwordBytes, salt)

                self.c.execute("INSERT INTO admins (manager_ID, username, password) VALUES(?, ?, ?)",
                               (rows[0][0], username, hash))
                self.conn.commit()

                self.clearAddAdmin()

    def clearAddAdmin(self):
        self.employeeIDLE.setText("")
        self.usernameLE.setText("")
        self.passwordLE.setText("")
        self.confirmLE.setText("")

        # reset table
        self.switchPage(3)

    def removeAdminPage(self):
        # switch to page
        self.stackedWidget_2.setCurrentIndex(6)

        self.removeEmpLE.setText("")


    def removeAdmnButton(self):
        employeeID = self.removeAdminLE.text()

        if not employeeID:
            self.msgbox.setText("Please enter employee ID")
            self.msgbox.exec_()
        else:
            self.c.execute("""SELECT hr_managers.manager_ID FROM hr_managers 
                           INNER JOIN admins ON hr_managers.manager_ID = admins.manager_ID
                           WHERE employee_ID = ?""", (employeeID,))
            rows = self.c.fetchall()

            if not rows:
                self.msgbox.setText(f"There is no admin with id of {employeeID}")
                self.msgbox.exec_()
                self.removeAdminLE.setText("")
            else:
                self.c.execute("DELETE FROM admins WHERE manager_ID = ?", (rows[0][0],))
                self.conn.commit()

                self.removeAdminLE.setText("")

                # reset table
                self.switchPage(3)


    # send signal to previous window to return to that window
    def closeEvent(self, event):
        self.destroyed.emit()

        # release camera if left opened
        if self.capture is not None:
            if self.capture.isOpened():
                self.capture.release()
                # stop updating frames
                self.timer.stop()

    def showEvent(self, event):
        # call homepage to set up table
        self.switchPage(0)

    def exportAttendance(self):

        # create new excel workbook
        wb = Workbook()
        # get default sheet
        ws = wb.active
        ws.title = "Attendance Report"

        # get the date last monday
        today = datetime.date.today()
        last_monday = today + rdelta.relativedelta(days=-1, weekday=rdelta.MO(-1))
        last_monday_str = str(last_monday)
        # get all attendance starting from last monday
        self.c.execute("""SELECT employees.employee_ID, employees.name, departments.name, date, time_in, time_out   
                        FROM employees 
                        INNER JOIN attendance ON employees.employee_ID = attendance.employee_ID
                        INNER JOIN departments ON employees.department_ID = departments.department_ID
                        WHERE date BETWEEN ? AND ?
                        """, (str(last_monday), str(today)))
        rows = self.c.fetchall()

        # add headings for excel table
        ws.append(["ID", "Name", "Department", "Date", "Time in", "Time out", "Total Time", "Undertime", "Overtime", "Tardy"])

        undertime = None
        overtime = None
        tardy = None

        for row in rows:
            row = list(row)
            if row[4] is not None:
                time_in = datetime.datetime.strptime(row[3] + " " + row[4], '%Y-%m-%d %H:%M:%S')
            else:
                row[4] = "None"

            if row[5] is not None:
                time_out = datetime.datetime.strptime(row[3] + " " + row[5], '%Y-%m-%d %H:%M:%S')
                elapsed_hours = time_out - time_in
                hours = abs(elapsed_hours.total_seconds() / 60 ** 2)
                minutes = abs(elapsed_hours.total_seconds() / 60) % 60
                total_time = "{:.0f}".format(hours) + "h" + "{:.0f}".format(minutes) + "m"

                if hours < 8:
                    undertime = "Yes"
                    overtime = "No"
                elif hours > 8:
                    undertime = "No"
                    overtime = "Yes"

                start_time = datetime.datetime.strptime(row[3] + " 08:00:00", '%Y-%m-%d %H:%M:%S')
                if time_in > start_time:
                    tardy = "Yes"
                elif time_in < start_time:
                    tardy = "No"
            else:
                row[5] = "None"
                total_time = "N/A"
                undertime = "N/A"
                overtime = "N/A"
                tardy = "N/A"


            ws.append(row + [total_time] + [undertime] + [overtime] + [tardy])

        for col in range(1,11):
            ws[get_column_letter(col) + '1'].font = Font(bold=True)

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 10

        # let user select where to save file using tkinter
        tkinter.Tk().withdraw()  # prevents an empty tkinter window from appearing

        folder_path = filedialog.asksaveasfilename(filetypes=[('excel file', '*.xlsx')],
                    defaultextension='.xlsx', initialfile=f"{last_monday} to {today}")

        if folder_path == "":
            return

        # save workbook
        try:
            wb.save(folder_path)
        except Exception:
            self.msgbox.setText("Error: The excel file for attendance is open. Please close before saving.")
            self.msgbox.exec_()

if __name__ == "__main__":
    # set up qt app
    app = QApplication(sys.argv)

    window = AdminWindow()
    window.show()

    # cleanly exits when x button is clicked
    sys.exit(app.exec_())